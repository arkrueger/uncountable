from recipe import Recipe
import openpyxl as op
import os

class Generator:
    def __init__(self):
        return
    
    def generate(self, name: str, recipes: list, destination: str) -> bool:
        # helper function
        def build_section(worksheet, section: Recipe.Section, start_row: int, start_col: int) -> int:
            try:
                row, col = start_row, start_col
                row += 1
                ### LEFTHAND HEADER (if applicable)
                header_offset = 0
                if section.lefthand:
                    header_offset = 1 # skip over the "step number" heading for the regular headers
                    worksheet.cell(row=row, column=1).value = section.lefthand
                    # fill in the "step number" column
                    # this could be done under an if statement in the other for loop, but I find it more readable here  
                    for i,v in enumerate(section.rows):
                        worksheet.cell(row=header_offset+row+i, column=1).value = section.rows[i]["number"]
                ### STANDARD HEADERS
                for i,v in enumerate(section.headers):
                    if len(section.headers) == header_offset+i: break # avoids going out of range
                    worksheet.cell(row=row, column=col+i).value = section.headers[header_offset+i] # caveat: bad code, assumes that the lefthand header will be the first element. See the format document for ideas on building this information into the format itself to make the code more robust.
                ### TABLE BODY
                row += 1 if any(section.headers) else 0 # don't increment if all headers are empty strings
                for i, v in enumerate(section.rows):
                    for j,c in enumerate(section.columns):
                        if len(section.columns) == header_offset+j: break
                        worksheet.cell(row=row+i, column=col+j).value = section.rows[i][section.columns[header_offset+j]]
                return row+i+2 # the next section will start 2 rows below the lowest row of this section
            except Exception as e:
                print("Something went wrong while building ", section.name, " section. ", e)
        
        # main function
        try:
            os.chdir(destination)
            workbook = op.Workbook()
            worksheet = workbook.active
            ### RECIPE NAMES (Set the recipe names for each recipe within the runsheet)
            lowest_row = 1 # keep track of the lowest row so that we start all recipes' next sections on the same row
            offset = 1 # column offset for the recipe names
            step = 2 # formulation names are placed in 2 cells merged together, so to reach the next empty cell we need to step by 2
            i = 1 # beginning column because spreadsheets are 1-indexed
            for r in recipes:
                worksheet.cell(row=1, column=i+offset).value = r.name
                worksheet.merge_cells(start_row=1, start_column=i+offset, end_row=1, end_column=i+offset+1) # 2 columns wide
                worksheet.cell(row=1, column=i+offset).alignment = op.styles.Alignment(horizontal='center')
                i += step
            # outdent

            ### SECTIONS
            current_row = 3 # we know that the first section title begins at row 3, always
            num_of_sections = 3 # currently, we know of three sections: metadata, steps, and measurements
            for i in range(num_of_sections):
                current_col = 2
                for j,r in enumerate(recipes):
                    section_lowest_row = build_section(worksheet, r.sections[i], start_row=current_row, start_col=current_col)
                    lowest_row = max(lowest_row, section_lowest_row)
                    current_col += step
                ### SECTION TITLE
                worksheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=current_col)
                worksheet.cell(row=current_row, column=1).alignment = op.styles.Alignment(horizontal='center')
                worksheet.cell(row=current_row, column=1).value = "*** "+r.sections[i].name.upper()+" ***"
                # once we're done with all recipes for the given section, move the current row down to the current row
                current_row = lowest_row

            # Adjust all column widths
            for i in range(current_col):
                col_letter = op.utils.get_column_letter(i+1)
                worksheet.column_dimensions[col_letter].width = 20
            # save the worksheet and return
            workbook.save(name+".xlsx")
            return True
        except Exception as e:
            print("Something went wrong while generating the runsheet. ", e)
            return False